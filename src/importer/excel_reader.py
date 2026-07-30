"""
Excel 数据读取模块

使用 openpyxl 流式读取 Excel 文件，处理大数据量场景（约 4 万行 × 692 列）。
支持逐行 yield 数据，避免一次性加载全部数据到内存。
"""

from typing import Any, Dict, Generator, List, Optional, Tuple

import openpyxl


# ── 已知的元数据列名 ────────────────────────────────────────
# 这些列不参与 FAI 测量列的匹配
METADATA_COLUMNS: set = {"Line", "Project", "Vendor", "Yield3", "SN", "Time"}

# ── 列名别名映射 ──────────────────────────────────────
# 不同产品线/产线可能用不同列名, 读入时统一为规范名
# canonical_name -> [aliases]  (匹配别名后改名为规范名)
COLUMN_ALIASES: dict = {
    "SN":      ["SN", "FG号", "序列号", "ProductSN", "SerialNumber"],
    "Time":    ["Time", "检测时间", "测量时间", "测试时间", "TestTime", "Timestamp"],
    "Project": ["Project", "项目", "项目代号", "ProductCode", "Project Code", "ProjectName", "Yield1"],
    "Vendor":  ["Vendor", "供应商", "Vendor Code", "VendorName", "Yield2"],
    "Line":    ["Line", "线体", "产线", "CFG", "LineName", "ProductionLine"],
}


def read_data_sheet(
    filepath: str,
    sheet_name: str = "Data",
) -> Tuple[List[str], Generator[List[Any], None, None], int]:
    """
    流式读取 Data Sheet 的数据。

    使用 openpyxl read_only 模式逐行读取，避免将整个文件加载到内存。
    自动过滤表头为空的尾随列，处理 #REF! 错误值。

    Args:
        filepath: Excel 文件路径
        sheet_name: Data Sheet 名称，默认 "Data"

    Returns:
        (列名列表, 数据行生成器, 有效列数)
        列名列表：仅包含表头非空的列
        数据行生成器：每次 yield 一个列表，元素顺序与列名列表对应
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # ── 读取表头，记录有效列的索引 ───────────────────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    columns: List[str] = []
    valid_indices: List[int] = []

    seen_names: Dict[str, int] = {}
    for i, header in enumerate(header_row):
        if header is not None:
            name = str(header).strip()
            # 列名别名标准化: 把 FG号/检测时间 等映射为 SN/Time
            for canonical, aliases in COLUMN_ALIASES.items():
                if name in aliases:
                    name = canonical
                    break
            # 处理重复列名：给后续重复列添加 _N 后缀
            if name in seen_names:
                seen_names[name] += 1
                name = f"{name}_{seen_names[name]}"
            else:
                seen_names[name] = 0
            columns.append(name)
            valid_indices.append(i)

    total_cols = len(columns)

    # ── 数据行生成器 ─────────────────────────────────────
    def row_generator() -> Generator[List[Any], None, None]:
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                values: List[Any] = []
                for idx in valid_indices:
                    if idx >= len(row):
                        values.append(None)
                        continue
                    cell = row[idx]
                    # 处理 Excel 错误值 #REF!
                    if isinstance(cell, str) and "#REF!" in cell:
                        values.append(None)
                    else:
                        values.append(cell)
                yield values
        finally:
            wb.close()

    return columns, row_generator(), total_cols


def read_spec_sheet(
    filepath: str,
    sheet_name: str = "Spec",
) -> List[Dict[str, Any]]:
    """
    读取 Spec Sheet 的规格定义。

    每行定义了一个 FAI 测量项的上下限和标准值。

    Args:
        filepath: Excel 文件路径
        sheet_name: Spec Sheet 名称，默认 "Spec"

    Returns:
        spec 字典列表，每个字典包含:
        - fai_name: FAI 测量项名称
        - usl: 规格上限（999999 表示无上限）
        - nominal: 标准值
        - lsl: 规格下限（-999999 表示无下限）
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # ── 读取表头，建立列名到索引的映射 ──────────────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is not None:
            header_map[str(h).strip()] = i

    # 定位各列（处理可能的列名空格）
    fai_idx = header_map.get("FAI", 0)
    usl_idx = header_map.get("USL", 1)
    nom_idx = header_map.get("Nomial", 2)
    lsl_idx = header_map.get("LSL", 3)

    specs: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[fai_idx] is None:
            continue  # 跳过 FAI 名称为空的行

        fai_name = str(row[fai_idx]).strip()
        if not fai_name:
            continue

        spec: Dict[str, Any] = {
            "fai_name": fai_name,
            "usl": _safe_float(row[usl_idx] if usl_idx < len(row) else None),
            "nominal": _safe_float(row[nom_idx] if nom_idx < len(row) else None),
            "lsl": _safe_float(row[lsl_idx] if lsl_idx < len(row) else None),
        }
        specs.append(spec)

    wb.close()
    return specs


def detect_fai_columns(
    data_columns: List[str],
    spec_fai_names: List[str],
) -> Tuple[List[str], List[str]]:
    """
    从 Data Sheet 的列名中，匹配 Spec 中定义的 FAI 测量列。

    匹配规则：
    - 跳过已知的元数据列（Line, Project, Vendor, Yield3, SN, Time）
    - 精确匹配（忽略首尾空格）

    Args:
        data_columns: Data Sheet 的全部有效列名
        spec_fai_names: Spec Sheet 中定义的 FAI 名称列表

    Returns:
        (匹配的 FAI 列名列表, 未匹配的列名列表)
        匹配列表：在 Spec 中有对应规格定义的列
        未匹配列表：不在 Spec 中的列（可能是额外数据列或命名差异）
    """
    spec_set = {name.strip() for name in spec_fai_names}
    meta_lower = {m.lower() for m in METADATA_COLUMNS}

    matched: List[str] = []
    unmatched: List[str] = []

    for col in data_columns:
        col_stripped = col.strip()
        # 跳过已知元数据列
        if col_stripped.lower() in meta_lower:
            continue
        if col_stripped in spec_set:
            matched.append(col_stripped)
        else:
            unmatched.append(col_stripped)

    return matched, unmatched


# ── 内部辅助函数 ─────────────────────────────────────────


def _safe_float(value: Any) -> Optional[float]:
    """安全转换为 float，失败时返回 None"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None
