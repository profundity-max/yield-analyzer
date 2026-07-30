"""
Spec 数据加载器

从 Excel 文件读取 FAI 规格数据（FAI名称、规格上限、标准值、规格下限），
并处理 Sentinel 值（±999999 表示无上限/无下限）。
"""

from typing import Any, Optional

import openpyxl

from src.config import SPEC_INF_VALUE


def load_spec_from_excel(filepath: str, sheet_name: str = "Spec") -> list[dict[str, Any]]:
    """
    从 Excel 文件的指定 Sheet 中加载规格数据。

    解析约定的 4 列结构：
      - FAI:    FAI 测量项名称
      - USL:    规格上限（Upper Spec Limit）
      - Nomial: 标准值（Nominal）
      - LSL:    规格下限（Lower Spec Limit）

    Sentinel 值（配置中的 SPEC_INF_VALUE，默认 999999 / -999999）
    会被转换为 None，表示该方向无限制。

    Args:
        filepath:   Excel 文件路径。
        sheet_name: Sheet 名称，默认 "Spec"。

    Returns:
        list[dict]: 每行一个规格记录，键为:
                    fai_name, lower_limit, upper_limit, nominal。

    Raises:
        FileNotFoundError: 文件不存在时抛出。
        ValueError:        Sheet 不存在或列结构不符合预期时抛出。
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"找不到 Excel 文件: {filepath}")

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' 不存在。可用的 Sheet: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    spec_data: list[dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        # 跳过空行
        if row is None or all(cell is None for cell in row):
            continue

        fai_raw = row[0] if len(row) > 0 else None
        usl_raw = row[1] if len(row) > 1 else None
        nom_raw = row[2] if len(row) > 2 else None
        lsl_raw = row[3] if len(row) > 3 else None

        # 跳过表头行（第一行通常是标题）
        if row_idx == 1:
            # 判断是否为表头：如果第一列的值为字符串 "FAI" 或包含 "FAI"
            if fai_raw and isinstance(fai_raw, str) and fai_raw.strip().upper() in ("FAI", "FAI_NAME"):
                continue

        # FAI 名称为空则跳过
        if fai_raw is None or (isinstance(fai_raw, str) and fai_raw.strip() == ""):
            continue

        fai_name = str(fai_raw).strip()

        def _parse_limit(value: Any) -> Optional[float]:
            """将单元格值转为 float，Sentinel 值转为 None。"""
            if value is None:
                return None
            try:
                num = float(value)
            except (ValueError, TypeError):
                return None
            # Sentinel 值检测：±999999 表示无限制
            if abs(num) >= SPEC_INF_VALUE:
                return None
            return num

        upper_limit = _parse_limit(usl_raw)
        nominal = _parse_limit(nom_raw)
        lower_limit = _parse_limit(lsl_raw)

        spec_data.append({
            "fai_name": fai_name,
            "lower_limit": lower_limit,
            "upper_limit": upper_limit,
            "nominal": nominal,
        })

    wb.close()
    return spec_data


def get_fai_name_list(spec_data: list[dict[str, Any]]) -> list[str]:
    """
    从规格数据中提取所有 FAI 名称列表（保持原始顺序）。

    Args:
        spec_data: load_spec_from_excel 的返回值。

    Returns:
        list[str]: FAI 名称列表。
    """
    return [item["fai_name"] for item in spec_data]


def validate_spec(spec_data: list[dict[str, Any]]) -> list[str]:
    """
    校验规格数据的合理性。

    校验项：
      1. FAI 名称不能为空。
      2. FAI 名称不能重复。
      3. 如果同时提供了上限和下限，上限必须大于等于下限。
      4. 如果提供了标准值，且同时有上下限，标准值需在区间内。

    Args:
        spec_data: load_spec_from_excel 的返回值。

    Returns:
        list[str]: 校验错误信息列表。空列表表示数据全部通过校验。
    """
    errors: list[str] = []
    seen_names: set[str] = set()

    for idx, item in enumerate(spec_data, start=1):
        fai_name = item.get("fai_name", "")
        lower = item.get("lower_limit")
        upper = item.get("upper_limit")
        nominal = item.get("nominal")

        # 1. FAI 名称非空
        if not fai_name:
            errors.append(f"第 {idx} 行: FAI 名称为空")
            continue

        # 2. 重复检查
        if fai_name in seen_names:
            errors.append(f"第 {idx} 行: FAI 名称 '{fai_name}' 重复")
        seen_names.add(fai_name)

        # 3. 上下限合理性
        if lower is not None and upper is not None and lower > upper:
            errors.append(
                f"第 {idx} 行 ({fai_name}): 下限 ({lower}) 大于上限 ({upper})"
            )

        # 4. 标准值在区间内
        if nominal is not None:
            if lower is not None and nominal < lower:
                errors.append(
                    f"第 {idx} 行 ({fai_name}): 标准值 ({nominal}) 低于下限 ({lower})"
                )
            if upper is not None and nominal > upper:
                errors.append(
                    f"第 {idx} 行 ({fai_name}): 标准值 ({nominal}) 高于上限 ({upper})"
                )

    return errors
