"""
分析结果导出模块

提供按日期导出每日良率 + 每日 TOP N 不良的功能。
支持两种数据视图：
  - 回归前（pre-regression）：原始 judged 数据，每条 SN 独立计数
  - 回归后（post-regression）：SN 去重，同一 SN 取最新记录、归入首次投产日

输出格式：Excel (.xlsx)，包含「每日良率」和「每日TOP不良」两个 Sheet。
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pyarrow.parquet as pq
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import DAY_CUTOFF_HOUR, EXPORTS_DIR, JUDGED_DIR
from src.db import get_connection
from src.aggregator.data_source import get_default_source
from src.aggregator.queries import (
    DAILY_TOP_DEFECTS_QUERY,
    DAILY_TOP_DEFECTS_REGRESSION_QUERY,
    DAILY_YIELD_QUERY,
    SN_REGRESSION_QUERY,
)

logger = logging.getLogger(__name__)

# ── Excel 样式常量 ──────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
_HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
_DATA_FONT = Font(name="Microsoft YaHei", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center")



def _get_result_columns() -> list[str]:
    """
    获取 judged Parquet 中所有 _result 列名（不含 overall_result）
    """
    judged_files = sorted(JUDGED_DIR.glob("judged_*.parquet"))
    if not judged_files:
        return []
    pf = pq.ParquetFile(str(judged_files[0]))
    cols = [
        field.name for field in pf.schema_arrow
        if field.name.endswith("_result") and field.name != "overall_result"
    ]
    pf.close()
    return cols



def _style_header(ws, num_cols: int) -> None:
    """给表头行设置样式"""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, num_cols: int) -> None:
    """给数据行设置样式"""
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            if col_idx == 1:
                cell.alignment = _CENTER_ALIGN
            else:
                cell.alignment = _LEFT_ALIGN


def _auto_width(ws, num_cols: int, max_width: int = 40) -> None:
    """自动调整列宽（取表头和数据中最长的）"""
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
        adjusted = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted, 10)


def _has_data() -> bool:
    """检查是否有 judged 数据"""
    judged_files = sorted(JUDGED_DIR.glob("judged_*.parquet"))
    return len(judged_files) > 0


def _write_yield_sheet(
    wb: Workbook,
    daily_yield_rows: list[dict],
    rectification_rows: list[dict] | None = None,
) -> None:
    """
    写入「每日良率」Sheet

    列 (无 rectification):
      日期 | 总数 | OK数 | NG数 | 良率%

    列 (有 rectification, 回归后报表专用, ADR-0001):
      日期 | 总数 | OK数 | NG数 | 良率% | 预计整形NG | 预计OK | 预计良率%
    """
    ws = wb.active
    ws.title = "每日良率"

    has_rect = bool(rectification_rows)
    headers = ["日期", "总数", "OK数", "NG数", "良率%"]
    if has_rect:
        headers += ["预计整形NG", "预计OK", "预计良率%"]

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 把 rectification_rows 转成 dict by day
    rect_map = {}
    if rectification_rows:
        for r in rectification_rows:
            rect_map[str(r.get("production_day", ""))[:10]] = r

    for row_idx, row in enumerate(daily_yield_rows, 2):
        ws.cell(row=row_idx, column=1, value=str(row.get("production_day", ""))[:10])
        ws.cell(row=row_idx, column=2, value=row.get("total", 0))
        ws.cell(row=row_idx, column=3, value=row.get("ok_count", 0))
        ws.cell(row=row_idx, column=4, value=row.get("ng_count", 0))
        ws.cell(row=row_idx, column=5, value=row.get("yield_pct", 0.0))
        if has_rect:
            day_key = str(row.get("production_day", ""))[:10]
            rect = rect_map.get(day_key, {})
            ws.cell(row=row_idx, column=6, value=rect.get("rectifiable_count", 0))
            ws.cell(row=row_idx, column=7, value=rect.get("saved_count", 0))
            ws.cell(row=row_idx, column=8, value=rect.get("yield_pct_post", 0.0))

    _style_header(ws, len(headers))
    _style_data_rows(ws, 2, len(daily_yield_rows) + 1, len(headers))
    _auto_width(ws, len(headers))


def _write_defects_sheet(wb: Workbook, daily_defect_rows: list[dict]) -> None:
    """
    写入「每日TOP不良」Sheet

    列：日期 | 排名 | FAI名称 | NG数 | 当日总数 | 不良率%
    """
    ws = wb.create_sheet(title="每日TOP不良")

    headers = ["日期", "排名", "FAI名称", "NG数", "当日总数", "不良率%"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(daily_defect_rows, 2):
        ws.cell(row=row_idx, column=1, value=str(row.get("production_day", ""))[:10])
        ws.cell(row=row_idx, column=2, value=row.get("rank", 0))
        ws.cell(row=row_idx, column=3, value=row.get("fai_name", ""))
        ws.cell(row=row_idx, column=4, value=row.get("ng_count", 0))
        ws.cell(row=row_idx, column=5, value=row.get("total", 0))
        ws.cell(row=row_idx, column=6, value=row.get("ng_rate_pct", 0.0))

    _style_header(ws, len(headers))
    _style_data_rows(ws, 2, len(daily_defect_rows) + 1, len(headers))
    _auto_width(ws, len(headers))


def _build_filepath(prefix: str) -> str:
    """构建带时间戳的输出文件路径"""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return str(EXPORTS_DIR / f"{prefix}_{ts}.xlsx")


# ═══════════════════════════════════════════════════════════════
# 公开 API
# ═══════════════════════════════════════════════════════════════


def _query_daily_data(
    cutoff_hour: int, 
    regression: bool, 
    project: Optional[str] = None,
    line: Optional[str] = None,
) -> tuple[list[dict], list[dict], list[str], list[dict]]:
    """
    查询每日良率和每日TOP不良数据（内部共用函数）

    Args:
        cutoff_hour: 日切点小时
        regression: True=回归后, False=回归前
        project: 可选 Project 筛选
        line: 可选 Line 筛选

    Returns:
        (daily_yield_rows, daily_defect_rows, result_columns)
    """
    if not _has_data():
        raise RuntimeError("没有 judged 数据，请先执行数据导入和判定")

    conn = get_connection()
    result_columns = _get_result_columns()
    if not result_columns:
        raise RuntimeError("judged 数据中没有 _result 列，请检查数据格式")
    
    # Build extra WHERE conditions for Project/Line filter
    extra_where_parts = []
    if project:
        safe_p = project.replace("'", "''")
        extra_where_parts.append('"Project" = \'' + safe_p + '\'')
    if line:
        safe_l = line.replace("'", "''")
        extra_where_parts.append('"Line" = \'' + safe_l + '\'')
    extra_where = " AND ".join(extra_where_parts) if extra_where_parts else ""

    if regression:
        yield_sql = get_default_source().read_sql(SN_REGRESSION_QUERY, cutoff_hour=cutoff_hour,
            extra_where=extra_where,)
        defect_sql = get_default_source().read_sql(DAILY_TOP_DEFECTS_REGRESSION_QUERY, cutoff_hour=cutoff_hour,
            top_n=9999,
            extra_where=extra_where,)
        daily_yield_rows = [
            {
                "production_day": str(row[0]),
                "total": row[1],
                "ok_count": row[2],
                "yield_pct": row[3],
                "ng_count": row[1] - row[2],
            }
            for row in conn.execute(yield_sql).fetchall()
        ]
        # 回归后报表加预计整形良率 (ADR-0001)
        from src.aggregator.regression import get_daily_rectification_yield
        rectification_rows = get_daily_rectification_yield(cutoff_hour=cutoff_hour)
    else:
        yield_sql = get_default_source().read_sql(DAILY_YIELD_QUERY, cutoff_hour=cutoff_hour,
            cfg_filter=None,
            extra_where=extra_where,)
        defect_sql = get_default_source().read_sql(DAILY_TOP_DEFECTS_QUERY, cutoff_hour=cutoff_hour,
            top_n=9999,
            extra_where=extra_where,)
        daily_yield_rows = [
            {
                "production_day": str(row[0]),
                "total": row[1],
                "ok_count": row[2],
                "ng_count": row[3],
                "yield_pct": row[4],
            }
            for row in conn.execute(yield_sql).fetchall()
        ]
        rectification_rows = []  # 回归前报表无整形

    daily_defect_rows = [
        {
            "production_day": str(row[0]),
            "fai_name": row[1],
            "ng_count": row[2],
            "total": row[3],
            "ng_rate_pct": row[4],
            "rank": row[5],
        }
        for row in conn.execute(defect_sql).fetchall()
    ]

    return daily_yield_rows, daily_defect_rows, result_columns, rectification_rows


def _build_workbook_bytes(
    daily_yield_rows: list[dict],
    daily_defect_rows: list[dict],
    rectification_rows: list[dict] | None = None,
) -> bytes:
    """构建 Excel 工作簿并返回 bytes"""
    wb = Workbook()
    _write_yield_sheet(wb, daily_yield_rows, rectification_rows)
    _write_defects_sheet(wb, daily_defect_rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def get_report_bytes_pre(
    top_n: int = 10,
    cutoff_hour: int = DAY_CUTOFF_HOUR,
    project: Optional[str] = None,
    line: Optional[str] = None,
) -> bytes:
    """
    生成回归前日报 Excel 并返回 bytes（用于 Dashboard 下载）

    Args:
        top_n: 每天导出的 TOP N 不良数
        cutoff_hour: 日切点小时
        project: 可选 Project 筛选
        line: 可选 Line 筛选

    Returns:
        Excel 文件 bytes
    """
    daily_yield, daily_defects, _, _ = _query_daily_data(cutoff_hour, regression=False, project=project, line=line)
    # Filter to top_n per day
    daily_defects = [d for d in daily_defects if d["rank"] <= top_n]
    return _build_workbook_bytes(daily_yield, daily_defects)


def get_report_bytes_post(
    top_n: int = 10,
    cutoff_hour: int = DAY_CUTOFF_HOUR,
    project: Optional[str] = None,
    line: Optional[str] = None,
) -> bytes:
    """
    生成回归后日报 Excel 并返回 bytes（用于 Dashboard 下载）

    Args:
        top_n: 每天导出的 TOP N 不良数
        cutoff_hour: 日切点小时
        project: 可选 Project 筛选
        line: 可选 Line 筛选

    Returns:
        Excel 文件 bytes
    """
    daily_yield, daily_defects, _, rectification = _query_daily_data(cutoff_hour, regression=True, project=project, line=line)
    # Filter to top_n per day
    daily_defects = [d for d in daily_defects if d["rank"] <= top_n]
    return _build_workbook_bytes(daily_yield, daily_defects, rectification)


def export_daily_report(
    output_path: Optional[str] = None,
    top_n: int = 10,
    cutoff_hour: int = DAY_CUTOFF_HOUR,
) -> str:
    """导出回归前每日良率 + 每日 TOP N 不良 → Excel（文件版）"""
    if output_path is None:
        output_path = _build_filepath("daily_report_pre_regression")
    daily_yield, daily_defects, _, _ = _query_daily_data(cutoff_hour, regression=False)
    daily_defects = [d for d in daily_defects if d["rank"] <= top_n]
    wb = Workbook()
    _write_yield_sheet(wb, daily_yield)
    _write_defects_sheet(wb, daily_defects)
    wb.save(output_path)
    logger.info("回归前日报: %s (%d 天良率, %d 条不良)", output_path, len(daily_yield), len(daily_defects))
    return output_path


def export_daily_report_regression(
    output_path: Optional[str] = None,
    top_n: int = 10,
    cutoff_hour: int = DAY_CUTOFF_HOUR,
) -> str:
    """导出回归后每日良率 + 每日 TOP N 不良 → Excel（文件版）"""
    if output_path is None:
        output_path = _build_filepath("daily_report_post_regression")
    daily_yield, daily_defects, _, rectification = _query_daily_data(cutoff_hour, regression=True)
    daily_defects = [d for d in daily_defects if d["rank"] <= top_n]
    wb = Workbook()
    _write_yield_sheet(wb, daily_yield, rectification)
    _write_defects_sheet(wb, daily_defects)
    wb.save(output_path)
    logger.info("回归后日报: %s (%d 天良率, %d 条不良)", output_path, len(daily_yield), len(daily_defects))
    return output_path


def export_both_reports(
    output_dir: Optional[str] = None,
    top_n: int = 10,
) -> tuple[str, str]:
    """
    同时导出回归前和回归后两份日报

    Args:
        output_dir: 输出目录（默认 data/exports/）
        top_n: 每天导出的 TOP N 不良数

    Returns:
        (pre_path, post_path) 两份文件的路径
    """
    pre_path = export_daily_report(
        output_path=None if output_dir is None
        else str(Path(output_dir) / f"daily_report_pre_regression_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
        top_n=top_n,
    )
    post_path = export_daily_report_regression(
        output_path=None if output_dir is None
        else str(Path(output_dir) / f"daily_report_post_regression_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
        top_n=top_n,
    )
    return pre_path, post_path
